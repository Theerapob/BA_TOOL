import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_C = {
    "topic_bg":   "C6EFCE",
    "raw_bg":     "FFF2CC",
    "avro_bg":    "E2EFDA",
    "detail_bg":  "BDD7EE",
    "col_hdr_bg": "2E75B6",
    "col_hdr_fg": "FFFFFF",
    "pk_bg":      "FFD966",
    "logical_bg": "FFFFC0",
    "row_odd":    "FFFFFF",
    "row_even":   "F2F2F2",
}


def _s(ws, row, col, value, bg=None, fg="000000", bold=False,
       align_h="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(name="Arial", size=10, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    c.border = _BORDER
    return c


def _write_raw_section(ws, table_name: str, columns: list, start_row: int) -> int:
    """
    Section 1 — Raw (SQL Server)
    NO. | Name | PK or Unique | Max Length | Format | Nullable | Description | Possible Value
    """
    # TABLE header
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, f"TABLE:    {table_name}",
       bg=_C["topic_bg"], bold=True, align_h="left")
    start_row += 1

    # Raw (SQL Server) label
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, "Raw (SQL Server)",
       bg=_C["raw_bg"], bold=True, align_h="left")
    start_row += 1

    # Detail Section
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, "Detail Section",
       bg=_C["detail_bg"], bold=True, align_h="left")
    start_row += 1

    # Column headers
    for c, h in enumerate(
        ["NO.", "Name", "PK or Unique", "Max Length",
         "Format", "Nullable", "Description", "Possible Value"], 1
    ):
        _s(ws, start_row, c, h,
           bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    start_row += 1

    # Data rows
    for i, col in enumerate(columns, 1):
        r = start_row
        bg = _C["row_odd"] if i % 2 == 1 else _C["row_even"]
        is_pk = "Y" if col.get("is_pk") else "N"

        sql_type = col.get("source_sql_type", "")
        m = re.search(r"\(([^)]+)\)", sql_type)
        max_len = m.group(1) if m else "-"

        base_type = re.split(r"[\(\s]", sql_type.lower().strip())[0]

        _s(ws, r, 1, i,           bg=bg)
        _s(ws, r, 2, col.get("column_name", ""),   bg=bg, align_h="left")
        _s(ws, r, 3, is_pk,       bg=_C["pk_bg"] if is_pk == "Y" else bg)
        _s(ws, r, 4, max_len,     bg=bg)
        _s(ws, r, 5, base_type,   bg=bg)
        _s(ws, r, 6, col.get("nullable", ""),      bg=bg)
        _s(ws, r, 7, "",          bg=bg, align_h="left", wrap=True)
        _s(ws, r, 8, "",          bg=bg, align_h="left", wrap=True)
        start_row += 1

    return start_row + 1  # blank gap


def _write_avro_section(ws, table_name: str, columns: list, start_row: int) -> int:
    """
    Section 2 — Confluent (AVRO)
    NO. | Name | Partition Key | Raw Format type | Logical Format type | direct move / logic | Possible Value
    """
    # Topic row
    ws.merge_cells(f"A{start_row}:G{start_row}")
    _s(ws, start_row, 1,
       f"Topic:    UAT_EEAS_RAW_dbWorkforce_{table_name}",
       bg=_C["topic_bg"], bold=True, align_h="left")
    start_row += 1

    # Confluent (AVRO) label
    ws.merge_cells(f"A{start_row}:G{start_row}")
    _s(ws, start_row, 1, "Confluent (AVRO)",
       bg=_C["avro_bg"], bold=True, align_h="left")
    start_row += 1

    # Detail Section
    ws.merge_cells(f"A{start_row}:G{start_row}")
    _s(ws, start_row, 1, "Detail Section",
       bg=_C["detail_bg"], bold=True, align_h="left")
    start_row += 1

    # Column headers
    for c, h in enumerate(
        ["NO.", "Name", "Partition Key", "Raw Format type",
         "Logical Format type", "direct move / logic", "Possible Value"], 1
    ):
        _s(ws, start_row, c, h,
           bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    start_row += 1

    # Data rows
    for i, col in enumerate(columns, 1):
        r = start_row
        bg = _C["row_odd"] if i % 2 == 1 else _C["row_even"]
        is_pk = "Y" if col.get("is_pk") else "N"

        _s(ws, r, 1, i,                              bg=bg)
        _s(ws, r, 2, col.get("column_name", ""),     bg=bg, align_h="left")
        _s(ws, r, 3, is_pk,                          bg=_C["pk_bg"] if is_pk == "Y" else bg)
        _s(ws, r, 4, col.get("raw_type", ""),        bg=bg)
        _s(ws, r, 5, col.get("logical_type", ""),    bg=_C["logical_bg"])
        _s(ws, r, 6, "Direct move",                  bg=bg)
        _s(ws, r, 7, "",                             bg=bg, align_h="left", wrap=True)
        start_row += 1

    return start_row + 1


def _build_sheet(ws, table_name: str, columns: list):
    """1 sheet = Raw section + AVRO section"""
    next_row = _write_raw_section(ws, table_name, columns, start_row=1)
    _write_avro_section(ws, table_name, columns, start_row=next_row)

    for col, w in zip("ABCDEFGH", [8, 22, 14, 17, 19, 18, 40, 40]):
        ws.column_dimensions[col].width = w


# ── Public API ────────────────────────────────────────────────────

def export_confluent_xlsx(tables: dict) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for table_name, columns in tables.items():
        ws = wb.create_sheet(title=table_name[:31])
        _build_sheet(ws, table_name, columns)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_table_xlsx(columns: list, table_name: str = "Sheet1") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    _build_sheet(ws, table_name, columns)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Legacy (ชื่อเดิม — main.py ไม่ต้องแก้) ───────────────────────

def export_all_xlsx(tables: dict) -> io.BytesIO:
    return export_confluent_xlsx(tables)


def _build_csv_rows(table_name: str, columns: list) -> list:
    import re
    rows = []

    # Raw section
    rows.append([f"TABLE:    {table_name}"])
    rows.append(["Raw (SQL Server)"])
    rows.append(["Detail Section"])
    rows.append(["NO.", "Name", "PK or Unique", "Max Length", "Format", "Nullable", "Description", "Possible Value"])
    for i, col in enumerate(columns, 1):
        sql_type = col.get("source_sql_type", "")
        m = re.search(r"\(([^)]+)\)", sql_type)
        max_len = m.group(1) if m else "-"
        base_type = re.split(r"[\(\s]", sql_type.lower().strip())[0]
        rows.append([i, col.get("column_name", ""), "Y" if col.get("is_pk") else "N",
                     max_len, base_type, col.get("nullable", ""), "", ""])

    rows.append([])  # blank gap

    # Confluent section
    rows.append([f"Topic:    UAT_EEAS_RAW_dbWorkforce_{table_name}"])
    rows.append(["Confluent (AVRO)"])
    rows.append(["Detail Section"])
    rows.append(["NO.", "Name", "Partition Key", "Raw Format type", "Logical Format type", "direct move / logic", "Possible Value"])
    for i, col in enumerate(columns, 1):
        rows.append([i, col.get("column_name", ""), "Y" if col.get("is_pk") else "N",
                     col.get("raw_type", ""), col.get("logical_type", ""), "Direct move", ""])

    return rows


def _csv_bytes(rows: list) -> io.BytesIO:
    import csv
    import codecs
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    output = io.BytesIO(codecs.BOM_UTF8 + buf.getvalue().encode("utf-8"))
    output.seek(0)
    return output


def export_all_csv(tables: dict) -> io.BytesIO:
    all_rows = []
    first = True
    for table_name, columns in tables.items():
        if not first:
            all_rows.append([])
        all_rows.extend(_build_csv_rows(table_name, columns))
        first = False
    return _csv_bytes(all_rows)


def export_table_csv(columns: list, table_name: str = "Sheet1") -> io.BytesIO:
    return _csv_bytes(_build_csv_rows(table_name, columns))